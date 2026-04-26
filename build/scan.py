import os
import openpyxl

def find_text_in_xlsx_files(folder_path, search_text):
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            workbook = openpyxl.load_workbook(file_path)
            for worksheet_name in workbook.sheetnames:
                worksheet = workbook[worksheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value and search_text in str(cell.value):
                            print(f"⚠️ {file_name} : {worksheet_name}--{cell.coordinate} : {cell.value}")

folder_path = r"C:\Users\heavy\AppData\Local\Interject\FileCache"  # Replace with the path to your folder
search_text = 'nisc-'     # Replace with the text you are searching for
find_text_in_xlsx_files(folder_path, search_text)